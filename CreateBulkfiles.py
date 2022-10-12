import PyPDF2
import os
import re
import openpyxl

your_excel = openpyxl.load_workbook('Output_from_pdf.xlsx')
your_sheet = your_excel['Sheet1']

for file_name in os.listdir('all_format_pdf'):
    print(file_name)
    load_pdf = open(r'C:\\Users\HP\PycharmProjects\Capstoneproject\\all_format_pdf\pdf1.pdf', 'rb')
    read_pdf = PyPDF2.PdfFileReader(load_pdf)
    page_count = read_pdf.getNumPages()
    first_page = read_pdf.getPage(0)
    page_content = first_page.extractText()
    page_content = page_content.replace('\n', '')
    print(page_content)

    mobile_number = re.search(r'(?:\+?\d{2}[ -]?)?\d{10}', page_content).group()
    print(mobile_number)

    pincode = re.search(r'(?:\+?\d{2}[ -]?)?\d{6}', page_content).group()
    print(pincode)

    policy_num = re.search(r'(?:\+?\d{2}[ -]?)?\d{5}', page_content).group()
    print(policy_num)

    found_name = re.search(r'(?<=Name: )(.*)(?=City: )', page_content).group()
    print(found_name)

    city = re.search(r'(?<=City: )(.*)(?=Pin: )', page_content).group()
    print(city)

    Place_of_accident = re.search(r"(?<=Placeof accident : )(.*)(?=Typeof loss: )", page_content).group()
    print(Place_of_accident)

    Date_of_accident = re.search(r"(?<=Date of accident : )(.*)(?=Placeof )", page_content).group()
    print(Date_of_accident)

    Police_report_filed = re.search(r'(?<=Police report filed: )(.*)(?=Witness present: )', page_content).group()
    print(Police_report_filed)

    witness_present = re.search(r'(?<=Witness present: )(.*)(?=Agent type: )', page_content).group()
    print(witness_present)

    Agent_type = re.search(r'(?<=Agent type: )(.*)(?=End )', page_content).group()
    print(Agent_type)

    Type_of_loss = re.search(r'(?<=Typeof loss: )(.*)(?=Submitted )', page_content).group()
    print(Type_of_loss)

    Fault = re.search(r'(?<=Fault: )(.*)(?=Police )', page_content).group()
    print(Fault)

    Vehicle_number = re.search(r'(?<=Vehicle Number: )(.*)(?=Name: )', page_content).group()
    print(Vehicle_number)

    email = re.search(r'(?<=Email id: )(.*)(?=Date of )', page_content).group()
    print(email)

    Claim_amount = re.search(r'(?<=Submitted amount : )(.*)(?=Fault: )', page_content).group()
    print(Claim_amount)

    last_row_number = your_sheet.max_row
    print(last_row_number)

    your_sheet.cell(column=1, row=last_row_number+1).value = policy_num
    your_sheet.cell(column=2, row=last_row_number + 1).value = Vehicle_number
    your_sheet.cell(column=3, row=last_row_number + 1).value = found_name
    your_sheet.cell(column=4, row=last_row_number + 1).value = city
    your_sheet.cell(column=5, row=last_row_number + 1).value = pincode
    your_sheet.cell(column=6, row=last_row_number + 1).value = mobile_number
    your_sheet.cell(column=7, row=last_row_number + 1).value = email
    your_sheet.cell(column=8, row=last_row_number + 1).value = Date_of_accident
    your_sheet.cell(column=9, row=last_row_number + 1).value = Place_of_accident
    your_sheet.cell(column=10, row=last_row_number + 1).value = Type_of_loss
    your_sheet.cell(column=11, row=last_row_number + 1).value = Claim_amount
    your_sheet.cell(column=12, row=last_row_number + 1).value = Fault
    your_sheet.cell(column=13, row=last_row_number + 1).value = Police_report_filed
    your_sheet.cell(column=14, row=last_row_number + 1).value = witness_present
    your_sheet.cell(column=15, row=last_row_number + 1).value = Agent_type

    your_excel.save('Output_from_pdf.xlsx')